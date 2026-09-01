"""Planilha de matriculas do Moonshot Club — a fonte precisa de contrato.

Diferente do controle de consultorias, aqui existem data de inicio, data de
termino e periodo do plano em formato de data completa. Nao ha ambiguidade a
resolver: o termino e declarado, nao inferido.

Cada aba e uma turma ("PB-ALPHAVILLE Agosto26", "PC-ALPHAVILLE Novembro25") ou
um recorte de controle (renovacao, cancelamentos, dados de venda).
"""
import datetime as dt
import re

import pandas as pd

from .texto import norm

# Abas que nao sao lista de matricula.
ABAS_IGNORAR = {'renovacao 2026', 'dados vendas novembro25 pc', 'alunos boleto nb',
                'pedidos de cancelamento'}

# Abas de controle: existem na planilha mas nao sao turmas. 'CANCELAMENTOS' e um
# apanhado de saidas; 'MOONSHOT ELITE' e nivel de plano — 33 das suas 45 alunas
# tambem aparecem na turma em que foram vendidas, e conta-la como turma
# duplicaria essas pessoas.
ABAS_CONTROLE_FIXAS = {'cancelamentos'}
LIMIAR_SOBREPOSICAO = 0.5  # acima disso a aba e recorte de outras, nao turma

COLUNAS = {
    'nome': ['nome', 'coluna 2', 'aluna'],
    'telefone': ['telefone'],
    'email': ['e-mail', 'email'],
    'cpf': ['cpf'],
    # Nas turmas antigas a coluna chama so 'Plano' e guarda o PERIODO ("12 MESES");
    # nas novas, 'Tipo de Plano' guarda o PRODUTO ("Moonshot Club"). Sao coisas
    # diferentes e precisam de apelidos separados.
    'produto': ['tipo de plano'],
    'consultor': ['consultor estrategico'],
    'cs': ['cs responsavel'],
    'contrato': ['contrato'],
    'status': ['status'],
    'status_financeiro': ['status financeiro'],
    'status_contrato': ['status contrato'],
    'inicio': ['data de inicio do plano'],
    'termino': ['data de termino do plano'],
    'periodo': ['periodo do plano', 'plano'],
    'pagamento': ['pagamento'],
    'responsavel_venda': ['responsavel venda'],
    'onboarding': ['onboarding'],
    'acesso': ['acesso plataforma'],
}
# 'plano' precisa vir depois de 'tipo de plano' para nao capturar a coluna errada;
# 'status' idem, senao engole 'status financeiro'.
ORDEM = ['nome', 'telefone', 'email', 'cpf', 'produto', 'consultor', 'cs', 'contrato',
         'status_financeiro', 'status_contrato', 'status', 'inicio', 'termino',
         'periodo', 'pagamento', 'responsavel_venda', 'onboarding', 'acesso']


def _cab(v):
    return re.sub(r'\s+', ' ', norm(v)).strip()


def _localizar_cabecalho(ws):
    for i in range(1, min(6, ws.max_row + 1)):
        linha = [_cab(ws.cell(i, j).value) for j in range(1, ws.max_column + 1)]
        if any(c in ('nome', 'coluna 2', 'aluna') for c in linha) and \
           any('data de inicio' in c or 'status' in c for c in linha):
            return i
    return None


def _mapear(ws, hi):
    linha = {j: _cab(ws.cell(hi, j).value) for j in range(1, ws.max_column + 1)}
    mapa, usados = {}, set()
    for apelido in ORDEM:
        for j, c in linha.items():
            if j in usados or not c or c == 'none':
                continue
            if any(c == v or c.startswith(v) for v in COLUNAS[apelido]):
                mapa[apelido] = j
                usados.add(j)
                break
    return mapa


def _periodo_meses(valor):
    """'12 MESES' -> 12. Devolve None quando nao declarado."""
    if valor is None:
        return None
    m = re.search(r'(\d{1,2})\s*m', norm(valor))
    return int(m.group(1)) if m else None


def carregar(caminho):
    """Uma linha por matricula, com inicio e termino declarados."""
    import openpyxl
    wb = openpyxl.load_workbook(caminho, data_only=True)
    linhas = []
    for aba in wb.sheetnames:
        if norm(aba) in ABAS_IGNORAR:
            continue
        ws = wb[aba]
        hi = _localizar_cabecalho(ws)
        if hi is None:
            continue
        mapa = _mapear(ws, hi)
        if 'nome' not in mapa:
            continue
        for i in range(hi + 1, ws.max_row + 1):
            nome = ws.cell(i, mapa['nome']).value
            if not nome or not str(nome).strip():
                continue
            reg = {'turma': aba.strip(), 'linha_planilha': i}
            for apelido, j in mapa.items():
                v = ws.cell(i, j).value
                if isinstance(v, dt.datetime):
                    reg[apelido] = v
                elif v is None or str(v).strip() == '':
                    reg[apelido] = None
                else:
                    reg[apelido] = re.sub(r'\s+', ' ', str(v)).strip()
            reg['periodo_meses'] = _periodo_meses(reg.get('periodo'))
            linhas.append(reg)
    d = pd.DataFrame(linhas)
    if not len(d):
        return d

    # Datas: so aceitamos o que veio como data de verdade da planilha.
    for c in ('inicio', 'termino'):
        d[c] = pd.to_datetime(d[c], errors='coerce')

    # Termino declarado e a verdade. Onde falta, calculamos pelo periodo — e
    # marcamos, para nao misturar dado com estimativa.
    calc = pd.Series(pd.to_datetime([
        (i + pd.DateOffset(months=int(p))) if pd.notna(i) else pd.NaT
        for i, p in zip(d['inicio'], d['periodo_meses'].fillna(12))]), index=d.index)
    d['termino_fonte'] = ['declarado' if pd.notna(t) else
                          ('calculado do periodo' if pd.notna(i) else 'sem data')
                          for t, i in zip(d['termino'], d['inicio'])]
    d['termino_efetivo'] = d['termino'].fillna(calc)

    d['status_norm'] = d['status'].map(_status)
    d['ativo'] = d['status_norm'] == 'ativo'
    return d


def _status(v):
    n = norm(v)
    if not n or n == 'none':
        return 'sem status'
    if 'cancel' in n:
        return 'cancelado'
    if 'encerr' in n:          # "ENCERROU PLANO" / "Encerrado": contrato ja terminou
        return 'encerrado'
    if 'pausad' in n:
        return 'pausado'
    if 'bloquead' in n:
        return 'bloqueado'
    if 'pendente' in n:
        return 'pendente de pagamento'
    if n.startswith('ativ'):
        return 'ativo'
    return n[:30]


# Status que tiram o contrato da fila de renovacao: ou ja saiu, ou ja acabou.
STATUS_FORA = {'cancelado', 'encerrado'}


def classificar_abas(d, limiar=LIMIAR_SOBREPOSICAO):
    """Separa turma de aba de controle.

    A regra e o proprio dado: se mais da metade das alunas de uma aba tambem
    aparece em outras abas, aquela aba e um recorte (nivel de plano, lista de
    cancelamento) e nao uma turma. Contar as duas somaria a mesma pessoa duas
    vezes num mes que talvez nem seja o do contrato dela.
    """
    if not len(d):
        return pd.DataFrame()
    x = d.copy()
    x['_k'] = x['nome'].map(_chave_nome)
    por = {t: set(g['_k']) for t, g in x.groupby('turma')}
    linhas = []
    for t, chaves in por.items():
        fora = set().union(*[v for u, v in por.items() if u != t]) if len(por) > 1 else set()
        sobre = len(chaves & fora) / len(chaves) if chaves else 0
        com_data = int(x.loc[x['turma'] == t, 'termino_efetivo'].notna().sum())
        if norm(t) in ABAS_CONTROLE_FIXAS:
            tipo, motivo = 'controle', 'aba de controle de cancelamento'
        elif not com_data:
            tipo, motivo = 'controle', 'nenhuma linha tem data de termino'
        elif sobre > limiar:
            tipo, motivo = ('controle',
                            f'{sobre:.0%} das alunas tambem aparecem em outras abas: '
                            f'e recorte, nao turma')
        else:
            tipo, motivo = 'turma', 'turma com datas proprias'
        linhas.append({'aba': t, 'tipo': tipo, 'alunas': len(chaves),
                       'com_data_de_termino': com_data,
                       'sobreposicao_com_outras_abas': round(100 * sobre, 1),
                       'motivo': motivo})
    return pd.DataFrame(linhas).sort_values(['tipo', 'alunas'], ascending=[True, False])


def por_turma(d, classificacao, inicio, meses=18):
    """Contagem turma a turma, sem canceladas, por mes de termino.

    E a leitura operacional: cada turma tem um numero de contratos chegando ao
    fim, e e assim que a renovacao e organizada. Nao deduplica entre turmas —
    quem renovou aparece na turma antiga e na nova, com dois contratos.
    """
    abas = set(classificacao.loc[classificacao['tipo'] == 'turma', 'aba'])
    v = d[d['turma'].isin(abas) & (~d['status_norm'].isin(STATUS_FORA))].copy()
    v = v.dropna(subset=['termino_efetivo'])
    v['mes'] = v['termino_efetivo'].dt.to_period('M')
    ini = pd.Period(inicio, freq='M')
    v = v[(v['mes'] >= ini) & (v['mes'] < ini + meses)]
    if not len(v):
        return pd.DataFrame(), pd.DataFrame()
    detalhe = (v.groupby(['turma', 'mes']).size().rename('nao_canceladas')
               .reset_index())
    detalhe['mes'] = detalhe['mes'].astype(str)
    total = d[d['turma'].isin(abas)].dropna(subset=['termino_efetivo']).copy()
    total['mes'] = total['termino_efetivo'].dt.to_period('M')
    total = total[(total['mes'] >= ini) & (total['mes'] < ini + meses)]
    conta = lambda sel: total[sel].groupby('mes').size()
    resumo_mes = pd.DataFrame({
        'no_total': total.groupby('mes').size(),
        'ativas': conta(total['status_norm'] == 'ativo'),
        'canceladas': conta(total['status_norm'] == 'cancelado'),
        'encerradas': conta(total['status_norm'] == 'encerrado'),
        'outros_status': conta(~total['status_norm'].isin(['ativo', 'cancelado', 'encerrado'])),
    }).fillna(0).astype(int)
    # Fica na fila de renovacao tudo que nao saiu nem acabou — inclui pendente de
    # pagamento e bloqueado, que ainda sao contrato vivo.
    resumo_mes['nao_canceladas'] = (resumo_mes['no_total'] - resumo_mes['canceladas']
                                    - resumo_mes['encerradas'])
    resumo_mes = resumo_mes.reset_index()
    resumo_mes['mes'] = resumo_mes['mes'].astype(str)
    # Mes sem nenhum contrato terminando nao informa nada na tabela; o grafico
    # ja mostra o vazio pelo espacamento.
    resumo_mes = resumo_mes[resumo_mes['nao_canceladas'] > 0].copy()
    resumo_mes['pct'] = (100 * resumo_mes['nao_canceladas']
                         / resumo_mes['nao_canceladas'].sum()).round(1)
    return resumo_mes, detalhe


def deduplicar(d):
    """Uma aluna pode aparecer em mais de uma turma (renovacao, migracao de
    plano). Fica a matricula de termino mais distante — e a que vale hoje."""
    if not len(d):
        return d
    d = d.copy()
    d['_chave'] = d['nome'].map(lambda x: re.sub(r'[^a-z ]', '', norm(x)).strip())
    d['_ord'] = d['termino_efetivo'].fillna(pd.Timestamp('1900-01-01'))
    return (d.sort_values(['_chave', '_ord'], ascending=[True, False])
            .drop_duplicates('_chave').drop(columns=['_chave', '_ord'])
            .sort_values('termino_efetivo'))


def projecao(d, inicio, meses=18, so_ativos=True):
    """Terminos por mes a partir de `inicio`, separando ativos de cancelados."""
    v = d.dropna(subset=['termino_efetivo']).copy()
    v['mes'] = v['termino_efetivo'].dt.to_period('M')
    ini = pd.Period(inicio, freq='M')
    janela = pd.period_range(ini, periods=meses, freq='M')
    v = v[v['mes'].isin(janela)]
    if not len(v):
        return pd.DataFrame()
    g = v.groupby('mes').apply(lambda x: pd.Series({
        'ativos': int((x['status_norm'] == 'ativo').sum()),
        'cancelados': int((x['status_norm'] == 'cancelado').sum()),
        'pausados': int((x['status_norm'] == 'pausado').sum()),
        'outros': int((~x['status_norm'].isin(['ativo', 'cancelado', 'pausado'])).sum()),
        'total_matriculas': len(x),
        'turmas': ', '.join(sorted(x['turma'].unique())[:3]),
    }), include_groups=False).reset_index()
    g['mes'] = g['mes'].astype(str)
    total_ativos = int((v['status_norm'] == 'ativo').sum())
    g['pct_dos_ativos'] = (100 * g['ativos'] / total_ativos).round(1) if total_ativos else 0
    return g


def resumo(d, inicio):
    v = d.dropna(subset=['termino_efetivo'])
    ini = pd.Period(inicio, freq='M')
    mes = v['termino_efetivo'].dt.to_period('M')
    linhas = [
        ('matriculas na planilha', len(d)),
        ('com data de inicio', int(d['inicio'].notna().sum())),
        ('com termino declarado', int((d['termino_fonte'] == 'declarado').sum())),
        ('termino calculado pelo periodo', int((d['termino_fonte'] == 'calculado do periodo').sum())),
        ('sem data alguma', int((d['termino_fonte'] == 'sem data').sum())),
        ('ativas', int((d['status_norm'] == 'ativo').sum())),
        ('canceladas', int((d['status_norm'] == 'cancelado').sum())),
        ('encerradas (plano ja terminou)', int((d['status_norm'] == 'encerrado').sum())),
        ('outros status', int((~d['status_norm'].isin(
            ['ativo', 'cancelado', 'encerrado'])).sum())),
        (f'terminam antes de {inicio}', int((mes < ini).sum())),
        (f'terminam de {inicio} em diante', int((mes >= ini).sum())),
        (f'ATIVAS terminando de {inicio} em diante',
         int(((mes >= ini) & (v['status_norm'] == 'ativo')).sum())),
    ]
    return pd.DataFrame(linhas, columns=['item', 'valor'])


def pedidos_cancelamento(caminho):
    """A aba de pedidos e um controle a parte e nem sempre esta refletida no
    status da matricula. A divergencia entre as duas e informacao."""
    import openpyxl
    wb = openpyxl.load_workbook(caminho, data_only=True)
    if 'PEDIDOS DE CANCELAMENTO' not in wb.sheetnames:
        return pd.DataFrame()
    ws = wb['PEDIDOS DE CANCELAMENTO']
    hi = None
    for i in range(1, min(6, ws.max_row + 1)):
        if norm(ws.cell(i, 1).value) == 'aluna':
            hi = i
            break
    if hi is None:
        return pd.DataFrame()
    cab = {norm(ws.cell(hi, j).value): j for j in range(1, ws.max_column + 1)
           if ws.cell(hi, j).value}
    linhas = []
    for i in range(hi + 1, ws.max_row + 1):
        nome = ws.cell(i, 1).value
        if not nome or not str(nome).strip():
            continue
        linhas.append({
            'nome': re.sub(r'\s+', ' ', str(nome)).strip(),
            'data_pedido': ws.cell(i, cab['data do pedido']).value if 'data do pedido' in cab else None,
            'consultor': ws.cell(i, cab['consultor']).value if 'consultor' in cab else None,
            'status_pedido': ws.cell(i, cab['status']).value if 'status' in cab else None,
        })
    return pd.DataFrame(linhas)


def _chave_nome(s):
    return re.sub(r'[^a-z ]', '', norm(s)).strip()


def conferir_cancelamentos(matriculas, pedidos):
    """Cruza pedidos de cancelamento com o status da matricula."""
    if not len(pedidos) or not len(matriculas):
        return pd.DataFrame(), pd.DataFrame()
    m = dict(zip(matriculas['nome'].map(_chave_nome), matriculas['status_norm']))
    p = pedidos.copy()
    p['status_na_matricula'] = p['nome'].map(_chave_nome).map(m)
    resumo_ = p['status_na_matricula'].fillna('nao encontrada na matricula').value_counts()
    divergentes = p[p['status_na_matricula'] == 'ativo']
    return (resumo_.rename_axis('status na matricula').rename('pedidos').reset_index(),
            divergentes)


def cruzar_com_base(matriculas, base, casar_fn):
    """Liga as matriculas a base do estudo, para saber quem termina e em que estado."""
    d = matriculas.rename(columns={'nome': 'nome_consultoria'})
    d = casar_fn(d, base)
    cols = ['id_aluna', 'em_risco', 'classe', 'fat_brl', 'score_oportunidade',
            'engajamento', 'produto_ancora', 'uf']
    return d.merge(base[[c for c in cols if c in base.columns]], on='id_aluna', how='left')


def projecao_detalhada(cruzado, inicio, meses=18, abas_turma=None):
    """Terminos por mes, sem canceladas, com o perfil de quem termina.

    Restrito as abas de turma: incluir as de controle contaria a mesma aluna
    duas vezes, em meses diferentes.
    """
    v = cruzado.dropna(subset=['termino_efetivo']).copy()
    if abas_turma is not None:
        v = v[v['turma'].isin(abas_turma)]
    v = v[~v['status_norm'].isin(STATUS_FORA)]
    v['mes'] = v['termino_efetivo'].dt.to_period('M')
    ini = pd.Period(inicio, freq='M')
    v = v[(v['mes'] >= ini) & (v['mes'] < ini + meses)]
    if not len(v):
        return pd.DataFrame()
    g = v.groupby('mes').apply(lambda x: pd.Series({
        'nao_canceladas': len(x),
        'na_base_do_estudo': int(x['id_aluna'].notna().sum()),
        'em_risco_hoje': int(x['em_risco'].fillna(False).sum()),
        'classe_A': int((x['classe'] == 'A').sum()),
        'fat_mediano': x['fat_brl'].median(),
        'turmas': ', '.join(sorted(x['turma'].unique())[:2]),
    }), include_groups=False).reset_index()
    g['mes'] = g['mes'].astype(str)
    g['pct_do_total'] = (100 * g['nao_canceladas'] / len(v)).round(1)
    return g


# Precedencia: uma aluna pode ter varias linhas de matricula (reentrou numa
# turma nova, esta em turma e no Elite). Quem cancelou em 2025 e voltou em 2026
# e aluna ativa; o inverso nao vale.
_ORDEM_STATUS = ['ativo', 'pendente de pagamento', 'bloqueado', 'encerrado', 'cancelado']

ROTULO_SITUACAO = {
    'ativa': 'Ativa',
    'em_cancelamento': 'Em processo de cancelamento',
    'finalizada': 'Finalizou o programa',
    'cancelada': 'Cancelada',
    'sem_matricula': 'Sem matrícula localizada',
}


def situacao_por_aluna(matriculas, base, casar_fn):
    """Situacao de contrato de cada aluna do estudo, das duas fontes.

    `status_cadastro` e o que a planilha de matriculas declara. `sinal_consultor`
    e o que o consultor relatou em texto livre. Os dois ficam em colunas
    separadas de proposito: um e registro, o outro e leitura de quem acompanha,
    e eles discordam com frequencia.

    `situacao_contrato` combina os dois com uma regra explicita:

    - cancelada: o cadastro diz cancelado
    - finalizada: o cadastro diz encerrado / encerrou plano
    - em_cancelamento: contrato vivo no cadastro, mas ou esta inadimplente
      (pendente de pagamento, bloqueado) ou o consultor relatou pedido de saida
      ou ausencia de contato. Nao e cancelamento; e o caminho ate ele.
    - ativa: contrato vivo e sem sinal de saida
    - sem_matricula: nao encontrada no cadastro. Nao quer dizer que saiu —
      quer dizer que e do Pro, ou que o nome nao casou.
    """
    d = matriculas.rename(columns={'nome': 'nome_consultoria'})
    d = casar_fn(d, base)
    d = d.dropna(subset=['id_aluna'])
    if not len(d):
        return pd.DataFrame(columns=['id_aluna', 'status_cadastro', 'situacao_contrato'])
    ordem = {s: i for i, s in enumerate(_ORDEM_STATUS)}
    d['_ord'] = d['status_norm'].map(ordem).fillna(len(ordem))
    melhor = d.sort_values('_ord').groupby('id_aluna').first()

    sinal = base.set_index('id_aluna')['engajamento'] if 'engajamento' in base.columns \
        else pd.Series(dtype=object)
    saida = {'risco_saida', 'sem_contato'}
    linhas = []
    for ident, r in melhor.iterrows():
        st = r['status_norm']
        eng = sinal.get(ident) if len(sinal) else None
        if st == 'cancelado':
            sit = 'cancelada'
        elif st == 'encerrado':
            sit = 'finalizada'
        elif st in ('pendente de pagamento', 'bloqueado') or (eng in saida):
            sit = 'em_cancelamento'
        else:
            sit = 'ativa'
        linhas.append({'id_aluna': ident, 'status_cadastro': st,
                       'sinal_consultor': eng, 'situacao_contrato': sit,
                       'turma_matricula': r.get('turma')})
    return pd.DataFrame(linhas)


def tabela_situacao(base):
    """Quantas alunas do estudo em cada situacao, e o que elas valem."""
    d = base.copy()
    d['situacao_contrato'] = d['situacao_contrato'].fillna('sem_matricula')
    g = d.groupby('situacao_contrato').agg(
        alunas=('id_aluna', 'size'),
        classe_A=('classe', lambda s: int((s == 'A').sum())),
        score_mediano=('score_oportunidade', 'median'),
        fat_mediano=('fat_brl', 'median')).reset_index()
    g['pct'] = (100 * g['alunas'] / len(d)).round(1)
    g['rotulo'] = g['situacao_contrato'].map(ROTULO_SITUACAO)
    return g.sort_values('alunas', ascending=False)
