"""Data de entrada, saida prevista (entrada + 12 meses) e projecao de termino.

A planilha nao tem uma data de entrada confiavel: ela mistura tres convencoes,
e em parte das linhas duas leituras sao possiveis com um ano de diferenca.
Este modulo resolve o que da para resolver, marca o resto, e devolve a
confianca de cada linha para que a projecao possa ser lida como faixa.
"""
import datetime as dt
import re

import pandas as pd

# Como o Excel exibe a celula revela o que a pessoa quis dizer ao digitar.
FMT_DATA_REAL = {'dd/MM/yyyy', 'dd/mm/yyyy', 'dd/mm', 'd/m', 'd/mmmm'}
FMT_MES_ANO = {'mmmm/yyyy', 'mmm/yyyy', 'mmm"/"yy'}
# 'mmm/d' exibe "nov/25": pode ser novembro DE 2025 (o "dia" e o ano) ou
# 25 de novembro. As duas leituras aparecem na planilha, em consultores
# diferentes, e so o contexto de cada linha desempata.
FMT_AMBIGUO = {'mmm/d', 'mmmm/d'}

JANELA_ENTRADA_CONSULTORIA = 200  # dias plausiveis entre entrar e a 1a consultoria


def _localizar(ws):
    """(linha_cabecalho, col_entrada, [cols_consultoria], col_aluna)."""
    hi = None
    for i in range(1, min(9, ws.max_row + 1)):
        if any('aluna' in str(ws.cell(i, j).value or '').strip().lower()
               for j in range(1, ws.max_column + 1)):
            hi = i
            break
    if hi is None:
        return None, None, [], None
    entrada, cons, aluna = None, [], None
    for j in range(1, ws.max_column + 1):
        n = str(ws.cell(hi, j).value or '').strip().lower()
        if not n:
            continue
        if n.startswith('aluna') and aluna is None:
            aluna = j
        elif (n.startswith('entrada') or n.startswith('data in')) and entrada is None:
            entrada = j
        elif re.search(r'consult', n):
            cons.append(j)
    return hi, entrada, cons, aluna


def _resolver(celula, ancora, hoje):
    """Devolve (data_entrada, confianca, regra)."""
    v, fmt = celula.value, celula.number_format
    if not isinstance(v, dt.datetime):
        return None, None, 'sem data'
    if fmt in FMT_DATA_REAL:
        return v, 'alta', f'formato de data completa ({fmt})'
    if fmt in FMT_MES_ANO:
        return dt.datetime(v.year, v.month, 1), 'alta', f'formato mes/ano ({fmt})'
    if fmt not in FMT_AMBIGUO:
        return v, 'media', f'formato nao catalogado ({fmt})'

    literal = v
    dia_ano = dt.datetime(2000 + v.day, v.month, 1)  # "nov/25" -> novembro de 2025

    # Regra que decide quase tudo: ninguem entra no programa no futuro. Onde as
    # duas leituras diferem por um ano, uma delas cai depois de hoje e morre.
    candidatos = [(literal, 'literal'), (dia_ano, 'mes/ano no "dia"')]
    passado = [(d, q) for d, q in candidatos if d <= hoje]
    if len(passado) == 1:
        d, q = passado[0]
        return d, 'alta', f'{q}: a outra leitura cairia no futuro'
    if not passado:
        return None, 'nao resolvida', 'as duas leituras caem no futuro'

    # As duas caem no mesmo mes: qual escolher e irrelevante para uma projecao
    # mensal, entao o mes e certo mesmo com a leitura ambigua.
    if (literal.year, literal.month) == (dia_ano.year, dia_ano.month):
        return literal, 'alta', 'as duas leituras caem no mesmo mes'

    if ancora:
        cabe = lambda d: d <= ancora and (ancora - d).days <= JANELA_ENTRADA_CONSULTORIA
        okL, okD = cabe(literal), cabe(dia_ano)
        if okL and not okD:
            return literal, 'alta', 'literal: unica leitura que antecede a 1a consultoria'
        if okD and not okL:
            return dia_ano, 'alta', 'mes/ano no "dia": unica leitura que antecede a 1a consultoria'
        if okL and okD:
            # As duas cabem. Fica a mais proxima da 1a consultoria.
            escolha = literal if (ancora - literal) <= (ancora - dia_ano) else dia_ano
            qual = 'literal' if escolha is literal else 'mes/ano no "dia"'
            return escolha, 'media', f'{qual}: as duas leituras cabem, fica a mais proxima da 1a consultoria'
        return None, 'nao resolvida', 'nenhuma leitura antecede a 1a consultoria'
    # Sem 1a consultoria para ancorar: sobra o teste de plausibilidade temporal.
    passL, passD = literal <= hoje, dia_ano <= hoje
    if passD and not passL:
        return dia_ano, 'media', 'sem ancora; so a leitura mes/ano cai no passado'
    if passL and not passD:
        return literal, 'media', 'sem ancora; so a leitura literal cai no passado'
    return None, 'nao resolvida', 'sem ancora e as duas leituras sao possiveis'


def carregar_datas(caminho, hoje=None):
    """Uma linha por aluna da planilha, com entrada resolvida e saida prevista."""
    import openpyxl
    hoje = hoje or dt.datetime.now()
    wb = openpyxl.load_workbook(caminho, data_only=True)
    linhas = []
    for aba in wb.sheetnames:
        ws = wb[aba]
        hi, ce, cons, ca = _localizar(ws)
        if hi is None or ce is None or ca is None:
            continue
        for i in range(hi + 1, ws.max_row + 1):
            nome = ws.cell(i, ca).value
            if not nome or str(nome).strip().lower().startswith('moonshot'):
                continue
            datas_cons = [ws.cell(i, j).value for j in cons
                          if isinstance(ws.cell(i, j).value, dt.datetime)
                          and ws.cell(i, j).number_format in FMT_DATA_REAL]
            ancora = min(datas_cons) if datas_cons else None
            entrada, conf, regra = _resolver(ws.cell(i, ce), ancora, hoje)
            linhas.append({
                'aba': aba.strip(),
                'nome_consultoria': re.sub(r'\s+', ' ', str(nome)).strip(),
                'entrada': entrada,
                'entrada_confianca': conf,
                'entrada_regra': regra,
                'entrada_bruta': ws.cell(i, ce).value,
                'entrada_formato': ws.cell(i, ce).number_format,
                'primeira_consultoria': ancora,
                'n_consultorias_datadas': len(datas_cons),
            })
    d = pd.DataFrame(linhas)
    if len(d):
        d['saida_prevista'] = d['entrada'].map(
            lambda x: (x + pd.DateOffset(months=12)) if pd.notna(x) else pd.NaT)
    return d


def projecao(datas, inicio, meses=18):
    """Quantos contratos vencem por mes a partir de `inicio`.

    Reporta em faixa: o cenario base usa so as entradas de confianca alta, e o
    cenario amplo soma as de confianca media. As nao resolvidas ficam de fora
    das duas e sao contadas a parte.
    """
    d = datas.dropna(subset=['saida_prevista']).copy()
    d['mes'] = d['saida_prevista'].dt.to_period('M')
    ini = pd.Period(inicio, freq='M')
    faixa = pd.period_range(ini, periods=meses, freq='M')
    alta = d[d['entrada_confianca'] == 'alta']
    media = d[d['entrada_confianca'] == 'media']
    linhas = []
    for p in faixa:
        a = int((alta['mes'] == p).sum())
        m = int((media['mes'] == p).sum())
        linhas.append({'mes': str(p), 'confianca_alta': a, 'mais_confianca_media': m,
                       'total_faixa': a + m})
    tab = pd.DataFrame(linhas)
    resumo = {
        'linhas na planilha': len(datas),
        'entrada resolvida (alta)': int((datas['entrada_confianca'] == 'alta').sum()),
        'entrada resolvida (media)': int((datas['entrada_confianca'] == 'media').sum()),
        'entrada nao resolvida': int((datas['entrada_confianca'] == 'nao resolvida').sum()),
        'sem data de entrada': int(datas['entrada'].isna().sum()
                                   - (datas['entrada_confianca'] == 'nao resolvida').sum()),
        f'vencem antes de {inicio}': int((d['mes'] < ini).sum()),
        f'vencem de {inicio} em diante': int((d['mes'] >= ini).sum()),
        'vencem depois da janela': int((d['mes'] > faixa[-1]).sum()),
    }
    return tab, pd.DataFrame([{'item': k, 'valor': v} for k, v in resumo.items()])


def cruzar_com_base(datas, base, casar_fn):
    """Liga a projecao a base do estudo para saber quem vence e em que estado."""
    d = casar_fn(datas, base)
    cols = ['id_aluna', 'em_risco', 'classe', 'fat_brl', 'score_oportunidade',
            'engajamento', 'produto_ancora', 'uf']
    m = d.merge(base[[c for c in cols if c in base.columns]], on='id_aluna', how='left')
    return m


def projecao_detalhada(cruzado, inicio, meses=18):
    """Projecao mes a mes, com o perfil de quem vence."""
    v = cruzado.dropna(subset=['saida_prevista']).copy()
    v['mes'] = pd.to_datetime(v['saida_prevista']).dt.to_period('M')
    ini = pd.Period(inicio, freq='M')
    v = v[(v['mes'] >= ini) & (v['mes'] < ini + meses)]
    if not len(v):
        return pd.DataFrame()
    g = v.groupby('mes').agg(
        contratos_vencendo=('nome_consultoria', 'size'),
        casadas_com_a_base=('id_aluna', 'count'),
        em_risco_hoje=('em_risco', 'sum'),
        classe_A=('classe', lambda s: int((s == 'A').sum())),
        fat_mediano=('fat_brl', 'median')).reset_index()
    g['mes'] = g['mes'].astype(str)
    g['pct_do_total'] = (100 * g['contratos_vencendo'] / len(v)).round(1)
    return g
